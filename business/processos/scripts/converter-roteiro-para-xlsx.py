import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

STATUS_OPTIONS = ["⬜ Pendente", "🔶 Em andamento", "✅ Concluído"]

def normalize_status(raw):
    raw = raw.strip()
    if not raw:
        return STATUS_OPTIONS[0]
    low = raw.lower()
    if 'conclu' in low or '✅' in raw:
        return STATUS_OPTIONS[2]
    if 'andamento' in low or '🔶' in raw:
        return STATUS_OPTIONS[1]
    return STATUS_OPTIONS[0]

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def convert(md_path, xlsx_path):
    with open(md_path, encoding='utf-8') as f:
        text = f.read()
    lines = text.splitlines()

    title_match = re.search(r'^#\s+(.*)$', text, re.MULTILINE)
    title = title_match.group(1) if title_match else "Roteiro"

    fase_match = re.search(r'\*\*Fase atual:\*\*\s*(.*)', text)
    mes_match = re.search(r'\*\*M[êe]s equivalente:\*\*\s*(.*)', text)

    # --- parse entregaveis table ---
    table_start = None
    for i, l in enumerate(lines):
        if l.strip().startswith('|') and 'Entreg' in l and 'Status' in l:
            table_start = i
            break
    entregaveis = []
    if table_start is not None:
        i = table_start + 2  # skip header + separator
        while i < len(lines) and lines[i].strip().startswith('|'):
            cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
            if len(cells) >= 3:
                entregaveis.append(cells[:3])
            i += 1

    # --- parse cronograma ---
    cronograma_idx = None
    for i, l in enumerate(lines):
        if re.match(r'^##\s+Cronograma', l.strip()):
            cronograma_idx = i
            break

    meses = []
    if cronograma_idx is not None:
        i = cronograma_idx + 1
        current = None
        while i < len(lines):
            l = lines[i].strip()
            m = re.match(r'^###\s+(.*)$', l)
            if m:
                if current:
                    meses.append(current)
                current = {"titulo": m.group(1), "atividades": [], "sessao": ""}
            elif l.startswith('-') and current is not None:
                item = l.lstrip('-').strip()
                item_clean = re.sub(r'\*\*|\*', '', item)
                if item_clean.startswith('★'):
                    current["sessao"] = item_clean.lstrip('★').strip()
                else:
                    current["atividades"].append(item_clean)
            i += 1
        if current:
            meses.append(current)

    wb = Workbook()

    # Sheet 1: Onde estou / Entregáveis
    ws1 = wb.active
    ws1.title = "Onde Estou"
    ws1["A1"] = title
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A2"] = f"Fase atual: {fase_match.group(1) if fase_match else ''}"
    ws1["A3"] = f"Mês equivalente: {mes_match.group(1) if mes_match else ''}"

    header_row = 5
    headers = ["Entregável", "Status", "Observação"]
    for c, h in enumerate(headers, start=1):
        ws1.cell(row=header_row, column=c, value=h)
    style_header(ws1, header_row, len(headers))

    r = header_row + 1
    for row in entregaveis:
        entregavel, status_raw, obs = row
        ws1.cell(row=r, column=1, value=entregavel)
        ws1.cell(row=r, column=2, value=normalize_status(status_raw))
        ws1.cell(row=r, column=3, value=obs)
        r += 1

    last_row = r - 1
    if last_row >= header_row + 1:
        dv = DataValidation(type="list", formula1='"{}"'.format(",".join(STATUS_OPTIONS)), allow_blank=True)
        ws1.add_data_validation(dv)
        dv.add(f"B{header_row+1}:B{last_row}")

    autosize(ws1, [45, 20, 55])
    for row in ws1.iter_rows(min_row=header_row+1, max_row=last_row, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 2: Cronograma
    ws2 = wb.create_sheet("Cronograma")
    headers2 = ["Mês", "Atividades", "Sessão marco"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    r = 2
    for mes in meses:
        ws2.cell(row=r, column=1, value=mes["titulo"])
        ws2.cell(row=r, column=2, value="\n".join(f"- {a}" for a in mes["atividades"]))
        ws2.cell(row=r, column=3, value=mes["sessao"])
        r += 1

    autosize(ws2, [35, 60, 45])
    for row in ws2.iter_rows(min_row=2, max_row=r-1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)
    print(f"OK: {xlsx_path} ({len(entregaveis)} entregaveis, {len(meses)} meses)")

if __name__ == '__main__':
    convert(sys.argv[1], sys.argv[2])
